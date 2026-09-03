# FORCE UPDATE V35 - RUN LOGS + ROLLING SAVE + ALL PREVIOUS FIXES
import requests
import time
import pandas as pd
import base64
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from io import BytesIO
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from supabase import create_client, Client
from itertools import groupby
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import API_LOGIN, API_PASSWORD, SUPABASE_URL, SUPABASE_KEY, EMAIL_SENDER, EMAIL_PASSWORD, EMAIL_RECEIVER

# --- CONNECT TO CLOUD ---
supabase = None
try:
    if SUPABASE_URL and SUPABASE_KEY:
        supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    else:
        print("⚠️ Supabase Credentials Missing in Secrets")
except Exception as e:
    print(f"Supabase Connection Error: {e}")

TARGET_DOMAIN = "edutap.in"
COMPETITORS = {
    "anujjindal": "anujjindal.in",
    "careerpower": "careerpower.in",
    "testbook": "testbook.com",
    "oliveboard": "oliveboard.in",
    "adda247": "adda247.com",
    "ixambee": "ixambee.com"
}

MAX_WORKERS = 5   # Lower to 3 if bulk inaccuracy still appears
SAVE_EVERY  = 50  # Flush results to Supabase after every N keywords

# ─────────────────────────────────────────────
# SUPABASE TABLE — run this SQL ONCE in Supabase SQL editor:
#
# CREATE TABLE IF NOT EXISTS run_logs (
#   id          BIGSERIAL PRIMARY KEY,
#   run_id      TEXT NOT NULL,
#   run_type    TEXT DEFAULT 'manual',
#   logged_at   TEXT NOT NULL,
#   level       TEXT DEFAULT 'info',
#   keyword     TEXT,
#   exam        TEXT,
#   kw_type     TEXT,
#   rank        INT,
#   ranked_url  TEXT,
#   message     TEXT
# );
# ─────────────────────────────────────────────

# --- FETCHERS ---
def fetch_all_rows(table_name):
    if not supabase: return pd.DataFrame()
    all_rows = []
    start = 0; batch_size = 1000
    while True:
        try:
            response = supabase.table(table_name).select("*").range(start, start + batch_size - 1).execute()
            rows = response.data
            if not rows: break
            all_rows.extend(rows)
            if len(rows) < batch_size: break
            start += batch_size
        except: break
    return pd.DataFrame(all_rows)

def get_current_month_cost():
    if not supabase: return 0.0
    try:
        current_month = datetime.now().strftime("%Y-%m")
        res = supabase.table("update_logs").select("total_cost").ilike("run_date", f"{current_month}%").execute()
        return sum(item['total_cost'] for item in res.data)
    except: return 0.0

def get_live_usd_inr_rate():
    try:
        r = requests.get("https://api.exchangerate-api.com/v4/latest/USD", timeout=3)
        if r.status_code == 200: return float(r.json().get('rates', {}).get('INR', 90.0))
    except: pass
    return 90.0

def get_dataforseo_balance():
    """Fetch current DataForSEO account balance in USD."""
    try:
        url  = "https://api.dataforseo.com/v3/appendix/user_data"
        auth = "Basic " + base64.b64encode(f"{API_LOGIN}:{API_PASSWORD}".encode()).decode()
        headers = {'Authorization': auth, 'Content-Type': 'application/json'}
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            balance = data['tasks'][0]['result'][0]['money']['balance']
            return float(balance)
    except Exception as e:
        print(f"⚠️ Could not fetch DataForSEO balance: {e}")
    return None

def get_all_keywords(): return fetch_all_rows("keywords_master")

def fetch_run_ids(last_n=10):
    """Return list of the last N distinct run_ids, newest first."""
    if not supabase: return []
    try:
        res = (supabase.table("run_logs")
               .select("run_id, run_type, logged_at")
               .order("id", desc=True)
               .limit(10000)
               .execute())
        if not res.data: return []
        seen = {}
        for r in res.data:
            rid = r['run_id']
            if rid not in seen:
                seen[rid] = {"run_id": rid, "run_type": r.get("run_type", "manual"), "logged_at": r.get("logged_at", "")}
            if len(seen) >= last_n:
                break
        return list(seen.values())
    except Exception as e:
        print(f"⚠️ fetch_run_ids error: {e}")
        return []

def fetch_logs_for_run(run_id):
    """Fetch all log entries for a given run_id."""
    if not supabase: return pd.DataFrame()
    all_rows = []; start = 0; batch = 1000
    while True:
        try:
            res = (supabase.table("run_logs")
                   .select("*")
                   .eq("run_id", run_id)
                   .order("id")
                   .range(start, start + batch - 1)
                   .execute())
            rows = res.data
            if not rows: break
            all_rows.extend(rows)
            if len(rows) < batch: break
            start += batch
        except Exception as e:
            print(f"⚠️ fetch_logs_for_run error: {e}")
            break
    return pd.DataFrame(all_rows) if all_rows else pd.DataFrame()


# ─────────────────────────────────────────────
# INTERNAL LOG WRITER
# ─────────────────────────────────────────────
_log_buffer = []

def _log(run_id, run_type, level, message, keyword=None, exam=None, kw_type=None, rank=None, ranked_url=None):
    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
    _log_buffer.append({
        "run_id":     run_id,
        "run_type":   run_type,
        "logged_at":  ist_now.strftime("%Y-%m-%d %H:%M:%S"),
        "level":      level,
        "keyword":    keyword or "",
        "exam":       exam or "",
        "kw_type":    kw_type or "",
        "rank":       rank,
        "ranked_url": ranked_url or "",
        "message":    message
    })

def flush_logs():
    global _log_buffer
    if not _log_buffer or not supabase: return
    try:
        for i in range(0, len(_log_buffer), 500):
            supabase.table("run_logs").insert(_log_buffer[i:i+500]).execute()
    except Exception as e:
        print(f"⚠️ flush_logs error: {e}")
    _log_buffer = []


# --- DB MANAGEMENT ---
def add_keyword(exam, keyword, kw_type, cluster="", volume=0, target_url=""):
    if not supabase: return False, "Database not connected"
    try:
        if supabase.table("keywords_master").select("*").eq("keyword", keyword.strip()).execute().data:
            return False, f"Keyword '{keyword}' already exists in database."
        supabase.table("keywords_master").insert({
            "exam": exam, "keyword": keyword.strip(), "type": kw_type,
            "cluster": cluster, "volume": volume, "target_url": target_url
        }).execute()
        return True, "Success"
    except Exception as e: return False, str(e)

def delete_bulk_keywords(keyword_list):
    if not supabase: return
    try: supabase.table("keywords_master").delete().in_("keyword", keyword_list).execute()
    except: pass

def clear_master_database():
    if not supabase: return
    try: supabase.table("keywords_master").delete().gt("id", 0).execute()
    except: pass

def normalize_url(url):
    return str(url).lower().replace("https://", "").replace("http://", "").replace("www.", "").strip("/") if url else ""

def process_bulk_upload(uploaded_file, mode="append"):
    if not supabase: return False, "Database not connected"
    if mode == "replace_all": clear_master_database()
    try:
        xls = pd.ExcelFile(uploaded_file)
        if mode == "replace_exam":
            for sheet in xls.sheet_names:
                try: supabase.table("keywords_master").delete().eq("exam", sheet.strip()).execute()
                except: pass
        existing_kws = set()
        if mode != "replace_all":
            try:
                res = supabase.table("keywords_master").select("keyword").execute()
                existing_kws = {str(r['keyword']).lower().strip() for r in res.data}
            except: pass
        rows_to_insert = []
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            df.columns = [str(col).strip() for col in df.columns]
            current_exam = sheet.strip()
            for _, row in df.iterrows():
                cluster = str(row.get('Cluster', '')).strip()
                t_url = str(row.get('Page URLs', row.get('Page URL', row.get('Target URL', '')))).strip()
                if t_url.lower() == 'nan': t_url = ""
                if cluster.lower() == 'nan': cluster = ""
                p_kw = str(row.get('Primary Keyword', '')).strip()
                p_vol = row.get('Volume', 0)
                try: p_vol = int(float(str(p_vol).replace(',', '')))
                except: p_vol = 0
                if p_kw and p_kw.lower() != 'nan':
                    kw_clean = p_kw.lower()
                    if kw_clean not in existing_kws:
                        rows_to_insert.append({"exam": current_exam, "keyword": p_kw, "type": "Primary",
                                                "cluster": cluster, "volume": p_vol, "target_url": t_url})
                        existing_kws.add(kw_clean)
                sec_block = str(row.get('Secondary Keywords', ''))
                if sec_block and sec_block.lower() != 'nan':
                    sec_kws = [k.strip() for k in sec_block.split('\n') if k.strip()]
                    vol_block = str(row.get('Volume.1', row.get('Secondary Volume', '')))
                    sec_vols = []
                    if vol_block and vol_block.lower() != 'nan':
                        for v in vol_block.split('\n'):
                            try: sec_vols.append(int(float(str(v).replace(',', '').strip())))
                            except: sec_vols.append(0)
                    for i, s_kw in enumerate(sec_kws):
                        if not s_kw: continue
                        kw_clean = s_kw.lower()
                        s_vol = sec_vols[i] if i < len(sec_vols) else 0
                        if kw_clean not in existing_kws:
                            rows_to_insert.append({"exam": current_exam, "keyword": s_kw, "type": "Secondary",
                                                    "cluster": cluster, "volume": s_vol, "target_url": t_url})
                            existing_kws.add(kw_clean)
        if rows_to_insert:
            for i in range(0, len(rows_to_insert), 1000):
                supabase.table("keywords_master").insert(rows_to_insert[i:i+1000]).execute()
        return True, f"Success! Added {len(rows_to_insert)} new keywords. (Duplicates were ignored safely)."
    except Exception as e: return False, f"Error: {str(e)}"


# ─────────────────────────────────────────────
# SINGLE KEYWORD FETCHER
# Bug Fix: max_workers=5, retry sleep at top of loop
# Now also writes human-readable logs per keyword
# ─────────────────────────────────────────────
def fetch_rank_single(item, run_id, run_type):
    keyword    = item['keyword']
    target_url = item.get('target_url', '')
    url        = "https://api.dataforseo.com/v3/serp/google/organic/live/advanced"
    payload    = [{"keyword": keyword, "location_code": 2356, "language_code": "en",
                   "device": "mobile", "os": "android", "depth": 20}]
    auth       = "Basic " + base64.b64encode(f"{API_LOGIN}:{API_PASSWORD}".encode()).decode()
    headers    = {'Authorization': auth, 'Content-Type': 'application/json'}

    accumulated_cost = 0.0
    final_res = None

    for attempt in range(1, 3):
        if attempt == 2:
            time.sleep(1.5)  # Sleep BEFORE retry, not after success

        res_data = {
            "keyword": keyword, "exam": item['exam'], "type": item['type'],
            "rank": 101, "url": "No Data", "bucket": "B4 (>20)", "target_rank": 101, "cost": 0,
            "comp_ranks": {k: 101 for k in COMPETITORS.keys()},
            "comp_urls":  {k: ""  for k in COMPETITORS.keys()}
        }

        try:
            response  = requests.post(url, headers=headers, json=payload, timeout=30)
            data      = response.json()
            accumulated_cost += data.get('cost', 0)

            if response.status_code == 200:
                try:
                    tasks = data.get('tasks', [])
                    if not tasks:
                        msg = f"No tasks returned by DataForSEO for this keyword"
                        _log(run_id, run_type, "warning", msg, keyword=keyword, exam=item['exam'], kw_type=item['type'])
                        continue

                    result = tasks[0].get('result')
                    if not result:
                        task_msg = tasks[0].get('status_message', 'Unknown error')
                        msg = f"DataForSEO task failed — reason: {task_msg}"
                        _log(run_id, run_type, "error", msg, keyword=keyword, exam=item['exam'], kw_type=item['type'])
                        continue

                    items_list = result[0].get('items', [])
                    best = 101; best_url = "Not Ranked"; target_f = 101
                    clean_t = normalize_url(target_url)
                    comp_found      = {k: 101 for k in COMPETITORS.keys()}
                    comp_urls_found = {k: ""  for k in COMPETITORS.keys()}

                    for item_res in items_list:
                        if item_res.get('type') == 'organic':
                            r_url   = item_res.get('url', '')
                            clean_r = normalize_url(r_url)
                            grp     = item_res['rank_group']
                            if TARGET_DOMAIN in r_url:
                                if grp < best: best, best_url = grp, r_url
                                if clean_t and clean_t in clean_r:
                                    if grp < target_f: target_f = grp
                            for c_key, c_domain in COMPETITORS.items():
                                if c_domain in r_url and grp < comp_found[c_key]:
                                    comp_found[c_key] = grp
                                    comp_urls_found[c_key] = r_url

                    bucket = "B4 (>20)"
                    if best <= 3: bucket = "B1 (1-3)"
                    elif best <= 10: bucket = "B2 (4-10)"
                    elif best <= 20: bucket = "B3 (11-20)"

                    res_data.update({'rank': best, 'url': best_url, 'bucket': bucket,
                                     'target_rank': target_f, 'comp_ranks': comp_found, 'comp_urls': comp_urls_found})

                    # ── Human-readable log ────────────────────────────────────
                    if best <= 20:
                        url_short = best_url[:80] + "..." if len(best_url) > 80 else best_url
                        msg = f"Ranked #{best} | Bucket: {bucket} | URL: {url_short}"
                        _log(run_id, run_type, "success", msg, keyword=keyword,
                             exam=item['exam'], kw_type=item['type'],
                             rank=best, ranked_url=best_url)
                    else:
                        msg = "Not in Top 20 — EduTap.in not found in first 20 Google results"
                        _log(run_id, run_type, "info", msg, keyword=keyword,
                             exam=item['exam'], kw_type=item['type'])

                except Exception as parse_err:
                    msg = f"Failed to read API response — technical error: {parse_err}"
                    _log(run_id, run_type, "error", msg, keyword=keyword, exam=item['exam'], kw_type=item['type'])

            else:
                err_msg = data.get('status_message', str(response.status_code))
                if 'balance' in err_msg.lower() or response.status_code == 402:
                    msg = f"⚠️ LOW BALANCE — DataForSEO rejected this keyword. Check your DataForSEO account balance. Details: {err_msg}"
                    level = "error"
                else:
                    msg = f"API returned an error (HTTP {response.status_code}): {err_msg}"
                    level = "error"
                print(f"❌ {keyword}: {msg}")
                _log(run_id, run_type, level, msg, keyword=keyword, exam=item['exam'], kw_type=item['type'])
                res_data['url'] = f"Err: {err_msg}"

        except Exception as e:
            msg = f"Network error — could not reach DataForSEO: {e}"
            print(f"❌ {keyword}: {msg}")
            _log(run_id, run_type, "error", msg, keyword=keyword, exam=item['exam'], kw_type=item['type'])
            res_data['url'] = f"Err: {str(e)}"

        if res_data['rank'] <= 20:
            res_data['cost'] = accumulated_cost
            return res_data
        final_res = res_data

    if final_res is None: final_res = res_data
    final_res['cost'] = accumulated_cost
    return final_res


# ─────────────────────────────────────────────
# PREV MAP — safe timestamp-filtered build
# Prevents prev/curr rank swap in email
# ─────────────────────────────────────────────
def build_prev_map_safe():
    if not supabase: return {}
    ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
    cutoff  = (ist_now - timedelta(minutes=2)).strftime("%Y-%m-%d %H:%M")
    all_rows = []; start = 0
    while True:
        try:
            response = (supabase.table("rankings")
                        .select("keyword, rank, date")
                        .lte("date", cutoff)
                        .range(start, start + 999)
                        .execute())
            rows = response.data
            if not rows: break
            all_rows.extend(rows)
            if len(rows) < 1000: break
            start += 1000
        except Exception as e:
            print(f"⚠️ build_prev_map_safe error: {e}")
            break
    if not all_rows: return {}
    df = pd.DataFrame(all_rows)
    df['date_dt'] = pd.to_datetime(df['date'])
    df = df.sort_values('date_dt')
    latest = df.groupby('keyword').tail(1)
    return {row['keyword']: int(row['rank']) for _, row in latest.iterrows()}


# ─────────────────────────────────────────────
# MAIN RUNNER
# Rolling save every 50 + run logging
# ─────────────────────────────────────────────
def perform_update(keywords_list, progress_bar=None, status_text=None, run_type="manual"):
    global _log_buffer
    _log_buffer = []

    ist_now  = datetime.utcnow() + timedelta(hours=5, minutes=30)
    date_str = ist_now.strftime("%Y-%m-%d %H:%M")
    run_id   = date_str

    total           = len(keywords_list)
    total_run_cost  = 0.0
    completed       = 0
    results_to_save = []
    pending_save    = []
    log_flush_count = 0

    _log(run_id, run_type, "info",
         f"Run started — {total} keywords queued | Workers: {MAX_WORKERS} | Started at: {date_str} IST")

    def flush_rankings(rows):
        if not rows or not supabase: return
        for i in range(0, len(rows), 500):
            try: supabase.table("rankings").insert(rows[i:i+500]).execute()
            except Exception as e: print(f"❌ Ranking batch save error: {e}")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {executor.submit(fetch_rank_single, item, run_id, run_type): item
                      for item in keywords_list}
        for future in as_completed(future_map):
            try:
                res = future.result()
            except Exception as e:
                item = future_map[future]
                msg = f"Unexpected crash for '{item.get('keyword','')}': {e}"
                print(f"❌ {msg}")
                _log(run_id, run_type, "error", msg,
                     keyword=item.get('keyword',''), exam=item.get('exam',''), kw_type=item.get('type',''))
                continue

            total_run_cost += res['cost']
            cr = res['comp_ranks']; cu = res['comp_urls']
            row = {
                "date": date_str, "keyword": res['keyword'], "exam": res['exam'], "type": res['type'],
                "rank": res['rank'], "url": res['url'], "bucket": res['bucket'], "target_rank": res['target_rank'],
                "rank_anujjindal": cr['anujjindal'], "rank_careerpower": cr['careerpower'],
                "rank_testbook":   cr['testbook'],   "rank_oliveboard":  cr['oliveboard'],
                "rank_adda247":    cr['adda247'],     "rank_ixambee":     cr['ixambee'],
                "url_anujjindal":  cu['anujjindal'],  "url_careerpower":  cu['careerpower'],
                "url_testbook":    cu['testbook'],    "url_oliveboard":   cu['oliveboard'],
                "url_adda247":     cu['adda247'],     "url_ixambee":      cu['ixambee']
            }
            results_to_save.append(row)
            pending_save.append(row)
            completed += 1
            log_flush_count += 1

            # Rolling save every SAVE_EVERY
            if len(pending_save) >= SAVE_EVERY:
                flush_rankings(pending_save)
                pending_save.clear()
                print(f"💾 Saved to DB: {completed}/{total}")

            # Rolling log flush every SAVE_EVERY
            if log_flush_count >= SAVE_EVERY:
                flush_logs()
                log_flush_count = 0

            if status_text: status_text.text(f"Processing... {completed}/{total}")
            if progress_bar: progress_bar.progress(completed / total)

    # Final flushes
    if pending_save:
        flush_rankings(pending_save)
        print(f"💾 Final ranking batch saved ({len(pending_save)} rows)")

    _log(run_id, run_type, "info",
         f"Run completed — {completed}/{total} keywords processed | "
         f"Total cost this run: ${total_run_cost:.4f}")
    flush_logs()

    try:
        if supabase:
            supabase.table("update_logs").insert({
                "run_date": date_str, "keywords_count": total, "total_cost": total_run_cost
            }).execute()
    except: pass

    return date_str, total_run_cost, results_to_save


# ─────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────
def send_email_alert(alerts_dict, subject_prefix="Automatic Run", all_checked_data=None, run_cost=None, dataforseo_balance=None):
    recipients = [e.strip() for e in EMAIL_RECEIVER.split(",")] if "," in EMAIL_RECEIVER else [EMAIL_RECEIVER]
    ist_now    = datetime.utcnow() + timedelta(hours=5, minutes=30)
    date_label = ist_now.strftime('%d %b %Y')
    has_alerts = any(alerts_dict.values())
    is_manual  = "Manual" in subject_prefix or "manual" in subject_prefix.lower()

    msg = MIMEMultipart()
    msg['From'] = EMAIL_SENDER
    msg['To']   = ", ".join(recipients)

    def fmt_rank(val):
        return "Not in Top 20" if val > 20 else val

    def generate_grouped_table(items_list):
        if not items_list: return ""
        items_list.sort(key=lambda x: x.get('exam', 'Others'))
        html = "<table border='1' cellpadding='5' style='border-collapse:collapse; width:100%; text-align:left;'>"
        for exam_name, group in groupby(items_list, key=lambda x: x.get('exam', 'Others')):
            html += f"<tr style='background-color:#2c3e50; color:white;'><th colspan='4' style='padding:8px; font-size:14px;'>{exam_name}</th></tr>"
            html += "<tr style='background-color:#ecf0f1;'><th>Type</th><th>Keyword</th><th>Previous Rank</th><th>Current Rank</th></tr>"
            for item in group:
                html += f"<tr><td>{item.get('type','-')}</td><td>{item['kw']}</td><td>{fmt_rank(item['prev'])}</td><td>{fmt_rank(item['curr'])}</td></tr>"
        html += "</table><br>"
        return html

    # ── Low balance urgent flag ───────────────────────────────────
    low_balance = dataforseo_balance is not None and dataforseo_balance < 5

    if has_alerts:
        subj_suffix = " ⚠️ TOP UP DATAFORSEO NOW" if low_balance else ""
        msg['Subject'] = f"{subject_prefix}: SEO Alert ({date_label}){subj_suffix}"
        html_body  = f"<h2>📉 {subject_prefix} Report ({date_label})</h2>"
        if low_balance:
            html_body += (
                f"<div style='background:#c0392b;color:white;padding:10px 14px;"
                f"border-radius:6px;font-size:14px;margin-bottom:12px;'>"
                f"🚨 <b>URGENT: DataForSEO balance is critically low (${dataforseo_balance:.4f}). "
                f"Please top up immediately — the next run will fail without sufficient balance.</b></div>"
            )
        html_body += "<p>Significant rank changes:</p>"
        if alerts_dict["red"]:
            html_body += "<h3 style='color:red;'>🔴 Critical: Dropped out of Top 10</h3>" + generate_grouped_table(alerts_dict["red"])
        if alerts_dict["orange"]:
            html_body += "<h3 style='color:orange;'>🟠 Warning: Dropped 4+ Positions</h3>" + generate_grouped_table(alerts_dict["orange"])
        if alerts_dict["yellow"]:
            html_body += "<h3 style='color:#b5b500;'>🟡 Alert: Dropped out of Top 3</h3>" + generate_grouped_table(alerts_dict["yellow"])
        if alerts_dict["green"]:
            html_body += "<h3 style='color:green;'>🟢 Celebration: Entered Top 3!</h3>" + generate_grouped_table(alerts_dict["green"])
    elif is_manual and all_checked_data:
        subj_suffix = " ⚠️ TOP UP DATAFORSEO NOW" if low_balance else ""
        msg['Subject'] = f"{subject_prefix}: Report Completed ({date_label}){subj_suffix}"
        html_body  = f"<h2>✅ Manual Run Completed ({date_label})</h2>"
        if low_balance:
            html_body += (
                f"<div style='background:#c0392b;color:white;padding:10px 14px;"
                f"border-radius:6px;font-size:14px;margin-bottom:12px;'>"
                f"🚨 <b>URGENT: DataForSEO balance is critically low (${dataforseo_balance:.4f}). "
                f"Please top up immediately — the next run will fail without sufficient balance.</b></div>"
            )
        html_body += "<p>No significant alerts. Full status:</p>"
        html_body += generate_grouped_table(all_checked_data)
    else:
        subj_suffix = " ⚠️ TOP UP DATAFORSEO NOW" if low_balance else ""
        msg['Subject'] = f"{subject_prefix}: All Stable ({date_label}){subj_suffix}"
        html_body  = f"<h2>✅ Automatic Update Completed ({date_label})</h2>"
        if low_balance:
            html_body += (
                f"<div style='background:#c0392b;color:white;padding:10px 14px;"
                f"border-radius:6px;font-size:14px;margin-bottom:12px;'>"
                f"🚨 <b>URGENT: DataForSEO balance is critically low (${dataforseo_balance:.4f}). "
                f"Please top up immediately — the next run will fail without sufficient balance.</b></div>"
            )
        html_body += "<p>No significant drops or changes detected. All keywords stable.</p>"

    # ── Cost & Balance footer (shown in every email) ──────────────
    footer_lines = []
    if run_cost is not None:
        footer_lines.append(f"<b>This run cost:</b> ${run_cost:.4f}")
    if dataforseo_balance is not None:
        if dataforseo_balance < 5:
            bal_color = "#c0392b"; bal_icon = "🔴"
        elif dataforseo_balance < 20:
            bal_color = "#e67e22"; bal_icon = "🟠"
        else:
            bal_color = "#27ae60"; bal_icon = "🟢"
        footer_lines.append(
            f"<b>DataForSEO balance:</b> "
            f"<span style='color:{bal_color};'>{bal_icon} ${dataforseo_balance:.4f}</span>"
        )
    if footer_lines:
        html_body += (
            "<br><hr style='border:1px solid #ddd;'>"
            "<div style='background:#f8f9fa;padding:10px 14px;border-radius:6px;"
            "font-size:13px;color:#555;margin-top:8px;'>"
            "💰 <b>Run Summary</b><br><br>"
            + "<br>".join(footer_lines)
            + "</div>"
        )

    msg.attach(MIMEText(html_body, 'html'))
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, recipients, msg.as_string())
        server.quit()
        print("📧 Email Sent Successfully!")
    except Exception as e:
        print(f"❌ Email failed: {e}")

# ─────────────────────────────────────────────
# WEEKLY AUTOMATIC EMAIL — XLSX ATTACHMENT ONLY
# Keeps the existing manual email function above unchanged.
# ─────────────────────────────────────────────
def _display_rank_for_excel(value):
    """Convert the internal 101 sentinel to a user-friendly Excel value."""
    try:
        rank = int(value)
        return rank if rank <= 20 else "Not in Top 20"
    except (TypeError, ValueError):
        return "No Result"


def _build_weekly_excel_report(master_df, results_data, run_date):
    """Build a formatted one-sheet XLSX containing the full current keyword dataset."""
    result_map = {
        str(row.get("keyword", "")).strip(): row
        for row in (results_data or [])
        if str(row.get("keyword", "")).strip()
    }

    try:
        run_dt = datetime.strptime(str(run_date), "%Y-%m-%d %H:%M")
    except (TypeError, ValueError):
        run_dt = datetime.utcnow() + timedelta(hours=5, minutes=30)

    date_label = run_dt.strftime("%d %b %Y")
    file_date = run_dt.strftime("%Y-%m-%d")
    filename = f"EduTap_SEO_Weekly_Report_{file_date}.xlsx"

    headers = [
        "Exam", "Cluster", "Keyword", "Type", "Volume", "Target URL",
        "Current Rank", "Ranked URL", "Target Rank", "Bucket",
        "Anuj Jindal Rank", "Anuj Jindal URL",
        "CareerPower Rank", "CareerPower URL",
        "Testbook Rank", "Testbook URL",
        "Oliveboard Rank", "Oliveboard URL",
        "Adda247 Rank", "Adda247 URL",
        "ixamBee Rank", "ixamBee URL",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "All Keywords"
    ws.sheet_view.showGridLines = False

    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "EduTap SEO Weekly Keyword Report"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0097A8")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Run Date: {date_label} | Total Keywords: {len(master_df)}"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 4
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    header_border = Border(bottom=Side(style="medium", color="0097A8"))

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    rank_columns = {7, 9, 11, 13, 15, 17, 19, 21}
    url_columns = {6, 8, 12, 14, 16, 18, 20, 22}

    light_row_fill = PatternFill("solid", fgColor="F7FBFC")
    top3_fill = PatternFill("solid", fgColor="E2F0D9")
    top10_fill = PatternFill("solid", fgColor="DDEBF7")
    top20_fill = PatternFill("solid", fgColor="FFF2CC")
    not_ranked_fill = PatternFill("solid", fgColor="FCE4D6")

    data_row = header_row + 1
    for _, master_row in master_df.iterrows():
        keyword = str(master_row.get("keyword", "")).strip()
        result = result_map.get(keyword, {})
        comp_ranks = result.get("comp_ranks", {}) or {}
        comp_urls = result.get("comp_urls", {}) or {}

        values = [
            master_row.get("exam", result.get("exam", "")),
            master_row.get("cluster", ""),
            keyword,
            master_row.get("type", result.get("type", "")),
            master_row.get("volume", 0),
            master_row.get("target_url", ""),
            _display_rank_for_excel(result.get("rank")),
            result.get("url", "No Result"),
            _display_rank_for_excel(result.get("target_rank")),
            result.get("bucket", "No Result"),
            _display_rank_for_excel(comp_ranks.get("anujjindal")),
            comp_urls.get("anujjindal", ""),
            _display_rank_for_excel(comp_ranks.get("careerpower")),
            comp_urls.get("careerpower", ""),
            _display_rank_for_excel(comp_ranks.get("testbook")),
            comp_urls.get("testbook", ""),
            _display_rank_for_excel(comp_ranks.get("oliveboard")),
            comp_urls.get("oliveboard", ""),
            _display_rank_for_excel(comp_ranks.get("adda247")),
            comp_urls.get("adda247", ""),
            _display_rank_for_excel(comp_ranks.get("ixambee")),
            comp_urls.get("ixambee", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            if pd.isna(value):
                value = ""
            elif hasattr(value, "item"):
                try:
                    value = value.item()
                except Exception:
                    pass

            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in url_columns or col_idx == 3))

            if data_row % 2 == 0:
                cell.fill = light_row_fill

            if col_idx in rank_columns:
                if isinstance(value, int):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if value <= 3:
                        cell.fill = top3_fill
                    elif value <= 10:
                        cell.fill = top10_fill
                    else:
                        cell.fill = top20_fill
                elif value in ("Not in Top 20", "No Result"):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = not_ranked_fill

            if col_idx in url_columns and isinstance(value, str) and value.startswith(("http://", "https://")):
                cell.hyperlink = value
                cell.style = "Hyperlink"
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        data_row += 1

    ws.freeze_panes = "A5"
    if data_row > header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:{last_col}{data_row - 1}"

    widths = {
        1: 18, 2: 24, 3: 38, 4: 12, 5: 11, 6: 42,
        7: 14, 8: 48, 9: 14, 10: 16,
        11: 16, 12: 42, 13: 18, 14: 42,
        15: 15, 16: 42, 17: 16, 18: 42,
        19: 15, 20: 42, 21: 15, 22: 42,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[header_row].height = 32

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), filename, date_label


def send_weekly_email_report(master_df, results_data, run_date, run_cost=None, dataforseo_balance=None):
    """
    Send the scheduled weekly email with only:
    1) date,
    2) formatted XLSX attachment containing all keywords,
    3) run summary / existing balance reminder.

    This function is intentionally separate from send_email_alert() so manual emails stay unchanged.
    """
    recipients = [e.strip() for e in EMAIL_RECEIVER.split(",")] if "," in EMAIL_RECEIVER else [EMAIL_RECEIVER]
    excel_bytes, filename, date_label = _build_weekly_excel_report(master_df, results_data, run_date)

    low_balance = dataforseo_balance is not None and dataforseo_balance < 5
    subj_suffix = " ⚠️ TOP UP DATAFORSEO NOW" if low_balance else ""

    msg = MIMEMultipart()
    msg["From"] = EMAIL_SENDER
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = f"📅 Weekly Automatic Run: SEO Report ({date_label}){subj_suffix}"

    html_body = f"<h2>📅 {date_label}</h2>"
    html_body += f"<p>📎 <b>Attached:</b> {filename}</p>"

    # Keep the existing low-balance reminder behavior for the weekly email.
    if low_balance:
        html_body += (
            f"<div style='background:#c0392b;color:white;padding:10px 14px;"
            f"border-radius:6px;font-size:14px;margin-bottom:12px;'>"
            f"🚨 <b>URGENT: DataForSEO balance is critically low (${dataforseo_balance:.4f}). "
            f"Please top up immediately — the next run will fail without sufficient balance.</b></div>"
        )

    summary_lines = [f"<b>Keywords checked:</b> {len(master_df)}"]
    if run_cost is not None:
        summary_lines.append(f"<b>This run cost:</b> ${run_cost:.4f}")

    if dataforseo_balance is not None:
        if dataforseo_balance < 5:
            bal_color = "#c0392b"; bal_icon = "🔴"
        elif dataforseo_balance < 20:
            bal_color = "#e67e22"; bal_icon = "🟠"
        else:
            bal_color = "#27ae60"; bal_icon = "🟢"
        summary_lines.append(
            f"<b>DataForSEO balance:</b> "
            f"<span style='color:{bal_color};'>{bal_icon} ${dataforseo_balance:.4f}</span>"
        )

    html_body += (
        "<hr style='border:1px solid #ddd;'>"
        "<div style='background:#f8f9fa;padding:10px 14px;border-radius:6px;"
        "font-size:13px;color:#555;margin-top:8px;'>"
        "💰 <b>Run Summary</b><br><br>"
        + "<br>".join(summary_lines)
        + "</div>"
    )

    msg.attach(MIMEText(html_body, "html"))

    attachment = MIMEApplication(
        excel_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    attachment.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(attachment)

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, recipients, msg.as_string())
        server.quit()
        print(f"📧 Weekly XLSX Email Sent Successfully! Attachment: {filename}")
    except Exception as e:
        print(f"❌ Weekly email failed: {e}")

